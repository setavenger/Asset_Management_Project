"""
This file cleans the Bloomberg data
Years were duplicated
monthly data was fine
"""
import openpyxl
from datetime import datetime
from openpyxl.utils.cell import get_column_letter
from util import validate_eoy


wb = openpyxl.load_workbook('data/Yearly_Monthly_Data.xlsx')

# Todo needs to be checked before running the program
ws_yearly = wb['Yearly']     # sheet name

block_height = 10   # number of rows per ticker
max_col = 'L'       # last column
row = 5             # start row (first ticker)
free_space = 1      # empty room between blocks

# days to keep 30.06.2012
first_day = datetime(2012, 6, 30)  # 30.06.2012

for company_index in range(0, 250):
    print()
    print('Currently on company: {}'.format(company_index + 1))
    column = 4
    deleted_counter = 0
    # delete all weekends (double dates / non-trading days)
    while True:

        if ws_yearly.cell(row=row, column=column).value is None:
            print(row, column)
            print('Deleted {} dates on company {}'.format(deleted_counter, company_index + 1))
            row += block_height + free_space
            break

        if ws_yearly.cell(row=row, column=column).value == datetime(2019, 12, 31):
            deleted_counter += 1
            for i in range(block_height + 1):
                ws_yearly.cell(row=row + i, column=column - 1).value = None
            # after deletion concatenate
            range_str = "{}{}:{}{}".format(get_column_letter(column), row, max_col, row + block_height)
            ws_yearly.move_range(range_str, cols=-1)

        if not validate_eoy(ws_yearly.cell(row=row, column=column).value):
            deleted_counter += 1
            for i in range(block_height + 1):
                ws_yearly.cell(row=row + i, column=column).value = None
            # after deletion concatenate
            range_str = "{}{}:{}{}".format(get_column_letter(column + 1), row, max_col, row + block_height)
            ws_yearly.move_range(range_str, cols=-1)
        column += 1

wb.save('data/ratio_data_clean.xlsx')
print('File Saved')
