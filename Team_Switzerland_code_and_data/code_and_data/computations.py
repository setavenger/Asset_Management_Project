"""
This module is for the major computations:
    - In the first part the data from our Bloomberg dataset is loaded into two pandas DataFrame objects
      we decided against using one-size fits all due to the different time intervals
        (1) total returns
        (2) R1-4 data (R5 is added in the percentile df)
    - We pulled more data from BB than we needed so during the process of extracting the data
      we also filtered via a blacklist of companies which did not meet our standards in regards to data sufficiency
    - This file contains:
        reading data
        calculating R1-5
        regression
        predictions
        minimum variance -> portfolio

    To run this code some folders might need to be created before hand
"""

import openpyxl
import pandas as pd
import numpy as np
import datetime
from util import try_float_iter, even_groups, build_date_list, momentum_dates
import math
import statsmodels.api as sm
from scipy.optimize import minimize, LinearConstraint

wb = openpyxl.load_workbook('data/Switzerland_vfinal_clean.xlsx')
wb_ratios = openpyxl.load_workbook('data/Yearly_Monthly_Data.xlsx')

# Todo needs to be checked before running the program
ws_balance_sheet = wb['balance_sheet_data']  # sheet name
ws_total_return = wb['total_returns']  # sheet name
ws_exclusions = wb['exclusions']  # sheet name

ws_yearly = wb_ratios['Yearly']  # sheet name

free_space = 1  # empty room between blocks
# !Hint! : removed 2020 only moving until 2019 Balance Sheet Data
years_obv = ['2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019']
risk_factors_all = ['R1', 'R2', 'R3', 'R4', 'R5']
# days to keep 30.06.2011
first_day = datetime.datetime(2011, 6, 30)  # 30.06.2011

company_blacklist = []
row_ticker = 2
while True:
    if not ws_exclusions.cell(row=row_ticker, column=1).value:
        break
    company_blacklist.append(ws_exclusions.cell(row=row_ticker, column=1).value)
    row_ticker += 1

row_tot_return = 4

df_total_returns = pd.DataFrame(dtype='float64', columns=[datetime.date(2020, 6, 30)] +
                                                         [ws_total_return.cell(row=row_tot_return,
                                                                               column=i).value.date()
                                                          for i in range(4, 110)])

for company_index in range(0, 203):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = f'{ws_total_return.cell(row=row_tot_return, column=1).value}'

    total_returns = np.array(list(try_float_iter(
        [ws_total_return.cell(row=row_tot_return + 1, column=i).value for i in range(3, 111)])))

    df_total_returns.loc[company_ticker] = [math.log(total_returns[i] / total_returns[i + 1])
                                            for i in range(len(total_returns) - 1)]

    # move to next company
    row_tot_return += 2 + free_space  # + x depends on the number of rows inserted

filtered = df_total_returns[~df_total_returns.index.isin(company_blacklist)]
df_total_returns = filtered.iloc[:100]

# !Hint! : balance sheet data from here
#  plan is to create single dfs per company and concat them via loc
#  this will create a MultiIndex
#  the columns will be matched in case a company misses data for a specific year

df_balance_sheet_data = pd.DataFrame(columns=years_obv)

frames = []
keys = []
row_bal = 4

frames_total = []
row_ratios = 5

# !Hint! : NESN first year balance ebit missing

for company_index in range(0, 203):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = f'{ws_balance_sheet.cell(row=row_bal, column=1).value}'
    if company_ticker in company_blacklist:
        row_bal += 12 + free_space  # + x depends on the number of rows inserted
        continue
    keys.append(company_ticker)
    columns_df = ['2019']
    column = 4
    while True:
        if ws_balance_sheet.cell(row=row_bal, column=column).value is None:
            break
            pass
        columns_df.append(str(ws_balance_sheet.cell(row=row_bal, column=column).value.year))
        column += 1

    df_company = pd.DataFrame(columns=columns_df)
    df_company_total = pd.DataFrame(columns=columns_df)

    # !Hint! : the company risk calculations and df
    projected_benefit_obligation = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 2, column=i).value
                                                                 for i in range(3, 12)])))
    fair_value_plan_assets = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 3, column=i).value
                                                           for i in range(3, 12)])))
    market_cap = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 4, column=i).value
                                               for i in range(3, 12)])))
    service_cost = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 6, column=i).value
                                                 for i in range(3, 12)])))
    interest_cost = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 7, column=i).value
                                                  for i in range(3, 12)])))
    pension_paid = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 8, column=i).value
                                                 for i in range(3, 12)])))
    ebit = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 9, column=i).value
                                         for i in range(3, 12)])))
    asset_allocation_debt = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 10, column=i).value
                                                          for i in range(3, 12)])))
    asset_allocation_stocks = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 11, column=i).value
                                                            for i in range(3, 12)])))

    # !HINT! : here come the ratios to add more data to the regression
    book_equity = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 2, column=i).value
                                                for i in range(3, 12)])))

    total_assets = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 3, column=i).value
                                                 for i in range(3, 12)])))

    debt_to_equity = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 4, column=i).value
                                                   for i in range(3, 12)])))

    sales_rev_turn = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 5, column=i).value
                                                   for i in range(3, 12)])))

    eps = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 6, column=i).value
                                        for i in range(3, 12)])))

    dvd_paid = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 7, column=i).value
                                             for i in range(3, 12)])))

    depr_amort = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 8, column=i).value
                                               for i in range(3, 12)])))

    net_income = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 9, column=i).value
                                               for i in range(3, 12)])))

    cash_from_ops = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_ratios + 10, column=i).value
                                                  for i in range(3, 12)])))

    pension_deficit = projected_benefit_obligation - fair_value_plan_assets

    r1 = pension_deficit / market_cap
    r2 = projected_benefit_obligation / market_cap
    r3 = (service_cost + interest_cost - pension_paid) / ebit
    r4 = asset_allocation_debt - asset_allocation_stocks

    df_company.loc['R1'] = r1[:len(columns_df)]
    df_company.loc['R2'] = r2[:len(columns_df)]
    df_company.loc['R3'] = r3[:len(columns_df)]
    df_company.loc['R4'] = r4[:len(columns_df)]

    frames.append(df_company)

    df_company_total.loc['R1'] = r1[:len(columns_df)]
    df_company_total.loc['R2'] = r2[:len(columns_df)]
    df_company_total.loc['R3'] = r3[:len(columns_df)]
    df_company_total.loc['R4'] = r4[:len(columns_df)]

    # !HINT! : calculate ratios || omitted
    market_to_book = market_cap / book_equity
    cash_to_dvd = cash_from_ops / dvd_paid
    pay_out_ratio = dvd_paid / eps

    # market_to_book = np.append(market_to_book[:len(columns_df)], np.NaN)
    # cash_to_dvd = np.append(cash_to_dvd[:len(columns_df)], np.NaN)
    # pay_out_ratio = np.append(pay_out_ratio[:len(columns_df)], np.NaN)
    # debt_to_equity = np.append(debt_to_equity[:len(columns_df)], np.NaN)

    df_company_total.loc['market_to_book'] = market_to_book[:len(columns_df)]
    # df_company_total.loc['cash_to_dvd'] = cash_to_dvd[:len(columns_df)]
    df_company_total.loc['pay_out_ratio'] = pay_out_ratio[:len(columns_df)]
    df_company_total.loc['debt_to_equity'] = debt_to_equity[:len(columns_df)]

    frames_total.append(df_company_total)
    # move to next company
    row_bal += 12 + free_space  # + x depends on the number of rows inserted
    row_ratios += 10 + free_space  # + x depends on the number of rows inserted


df_main_bal = pd.concat(frames, keys=keys)
df_main_bal_total = pd.concat(frames_total, keys=keys)

# !Hint! : only the top 100
keys = keys[:100]
df_main_bal = df_main_bal.loc[keys]  # 100 companies à 4 risks
df_main_bal_total = df_main_bal_total.loc[keys]

# !Hint! : from here on the quintiles will be created
df_group = pd.DataFrame(columns=years_obv[:-1], index=df_main_bal.index)
df_percentile = pd.DataFrame(columns=years_obv[:-1], index=df_main_bal.index)

for year in years_obv[:-1]:
    for risk_factor in ['R1', 'R2', 'R3', 'R4']:
        print(f'{year}-{risk_factor}')
        sub_df = df_main_bal.xs(risk_factor, level=1)[year].to_frame(f'{year}_{risk_factor}')
        sub_df['percentile'] = sub_df[f'{year}_{risk_factor}'].rank(pct=True)
        sub_df = sub_df.sort_values(ascending=True, by=f'{year}_{risk_factor}').dropna()
        group = even_groups(len(sub_df), 5, True)
        sub_df = sub_df.assign(group=group)
        sub_df_group = pd.concat([sub_df], keys=[risk_factor]).swaplevel()['group']
        sub_df_percentile = pd.concat([sub_df], keys=[risk_factor]).swaplevel()['percentile']

        # df_group
        for index in sub_df_group.index:
            df_group.loc[index][year] = sub_df_group.loc[index]
        # df_percentile
        for index in sub_df_group.index:
            df_percentile.loc[index][year] = sub_df_percentile.loc[index]

# !Hint! : R5 calculations
df_r5 = pd.DataFrame(columns=years_obv[:-1])
for comp in keys:
    df_r5.loc[comp] = df_percentile.loc[comp].mean()

for year in df_r5.columns:
    df = df_r5[year].sort_values(ascending=True).dropna()
    group = even_groups(len(df), 5, True)
    df = df.to_frame(year).assign(group=group)
    df['percentile'] = df[f'{year}'].rank(pct=True)
    for comp in df.index:
        df_percentile.loc[(comp, 'R5'), year] = df['percentile'].loc[comp]
        df_percentile = df_percentile.sort_index()

for year in years_obv[:-1]:
    print(f'R5-{year}')
    df = df_r5[year]
    df = df.sort_values(ascending=True).dropna()
    group = even_groups(len(df), 5, True)
    df = df.to_frame('R5_Percentile').assign(group=group)
    df_group_r5 = pd.concat([df], keys=['R5']).swaplevel()['group']
    for index in df_group_r5.index:
        df_group.loc[index, year] = df_group_r5.loc[index]

# !Hint! : create total returns for all quintile portfolios
total_return_dfs = {}
quartiles_only = {}
for year in years_obv[:-1]:
    columns_new = build_date_list(int(year), df_total_returns)
    risk_dfs = []
    df_quartiles_only = pd.DataFrame(index=pd.MultiIndex.from_product([risk_factors_all,
                                                                       [f'Q-{rank}' for rank in range(1, 6)]]),
                                     columns=columns_new)
    for risk in risk_factors_all:
        rank_dfs = []
        for rank in range(1, 6):
            print(f'Total Return {year}--{risk} Q-{rank}')
            indices = df_group.xs(risk, level=1)[year][(df_group.xs(risk, level=1)[year] == rank)].index
            df = df_total_returns.loc[indices][df_total_returns.columns.intersection(columns_new)]
            df_quartiles_only.loc[risk, f'Q-{rank}'] = df.mean()
            rank_dfs.append(df)

        rank_df = pd.concat(rank_dfs, keys=[f'Q-{i}' for i in range(1, 6)])
        risk_dfs.append(rank_df)

    df_for_year = pd.concat(risk_dfs, keys=risk_factors_all)

    total_return_dfs[year] = df_for_year
    df_quartiles_only['mean'] = df_quartiles_only.mean(axis=1)
    df_quartiles_only['std'] = df_quartiles_only.std(axis=1)
    quartiles_only[year] = df_quartiles_only

# !Hint! : add R5 to coefficient table
df_r5_mod = pd.concat([df_r5], keys=['R5']).swaplevel()
for year in years_obv[:-1]:
    for index in df_r5_mod.index:
        print(f"Adding R5 {index} {year}")
        df_main_bal.loc[index, year] = df_r5_mod.loc[index][year]
        df_main_bal_total.loc[index, year] = df_r5_mod.loc[index][year]

df_main_bal = df_main_bal.drop('2019', 1)
df_main_bal_total = df_main_bal_total.drop('2019', 1)

df_regression_data_raw = df_main_bal.copy()
min_variance_data = pd.DataFrame(columns=pd.MultiIndex.from_product([years_obv[:-1], ['Std. Dev.']]))
# !Hint! : add Momentum 1, 6, 12
for year in years_obv[:-1]:
    columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(int(year), df_total_returns)
    columns_return_year = build_date_list(int(year), df_total_returns)

    momentum_1 = df_total_returns[columns_momentum_1].apply(lambda x: np.exp(x))
    momentum_1['Momentum 1'] = momentum_1.product(axis=1) - 1

    momentum_6 = df_total_returns[columns_momentum_6].apply(lambda x: np.exp(x))
    momentum_6['Momentum 6'] = momentum_6.product(axis=1) - 1

    momentum_12 = df_total_returns[columns_momentum_12].apply(lambda x: np.exp(x))
    momentum_12['Momentum 12'] = momentum_12.product(axis=1) - 1

    return_year = df_total_returns[columns_return_year].apply(lambda x: np.exp(x))
    return_year['Return Year'] = return_year.product(axis=1) - 1

    min_variance_data[(year, 'Std. Dev.')] = df_total_returns[columns_momentum_12].std(axis=1)

    for index in df_total_returns.index:
        print(f"Adding Momentum {index} {year}")
        df_regression_data_raw.loc[(index, 'Momentum 1'), year] = momentum_1['Momentum 1'].loc[index]
        df_regression_data_raw.loc[(index, 'Momentum 6'), year] = momentum_6['Momentum 6'].loc[index]
        df_regression_data_raw.loc[(index, 'Momentum 12'), year] = momentum_12['Momentum 12'].loc[index]
        df_regression_data_raw.loc[(index, 'Return Year'), year] = return_year['Return Year'].loc[index]

df_regression_data_raw = df_regression_data_raw.sort_index()

# cols = ['Momentum 1', 'Momentum 12', 'Momentum 6', 'R1', 'R2', 'R3', 'R4', 'R5', 'Return Year']
cols = sorted(list(set(df_regression_data_raw.index.get_level_values(level=1))))
df_regression_data = pd.DataFrame(columns=cols)

for year in years_obv[:-1]:
    for company in keys:
        print(f'Transposing {year} {company}')
        row = pd.concat([df_regression_data_raw.loc[company, year].to_frame().transpose()],
                        keys=[company])
        df_regression_data = df_regression_data.append(row)

# !Hint! : Regression here
df_regression_data_clean = df_regression_data.dropna()
y = df_regression_data_clean['Return Year']
X = df_regression_data_clean.drop('Return Year', axis=1)
X = sm.add_constant(X)
# Note the difference in argument order
model = sm.OLS(y, X).fit()
# predictions = model.predict(X)  # make the predictions by the model


# Print out the statistics
print(model.summary())

# !HINT! : only significant values forecast
pvals = model.pvalues
pvals_sign = pvals[pvals < 0.05]
params_sign = model.params.loc[pvals_sign.index]
predictions = pd.Series(index=X.index, data=np.dot(X[pvals_sign.index].values, params_sign))

# set index to multilevel
predictions.index = pd.MultiIndex.from_tuples(predictions.index, names=('company', 'year'))

min_variance_portfolio_data = {}

for year in years_obv[:-1]:
    # !Hint! : Preprocessing
    top_20_expected_return = predictions.xs(year, level=1).sort_values(ascending=False)[:20]
    top_20_all_data = top_20_expected_return.to_frame('Expected Return').join(min_variance_data[(year, 'Std. Dev.')])
    top_20_all_data.columns = ['Expected Return', 'Std. Dev.']
    cov = df_total_returns.loc[top_20_expected_return.index].T.cov()

    columns = build_date_list(year=int(year), df=df_total_returns)
    # !Hint! : Optimization procedure
    w = 20 * [0.05]
    constraint = LinearConstraint(np.ones(len(w)), lb=1, ub=1)

    return_objective = top_20_expected_return.quantile(0.3)


    def target_return(w):
        return np.dot([w], np.array([top_20_all_data['Expected Return'].values])[0]) - return_objective


    def objective_function(weights, cov):
        return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


    cons = [constraint,
            {'type': 'eq', 'fun': target_return}]

    bounds = [(0.001, .15) for _ in w]
    res = minimize(
        objective_function,
        x0=[1 / len(w) for _ in w],
        args=(cov,),
        constraints=cons,
        bounds=bounds, )

    pf_weights = res.x
    top_20_monthly_returns = df_total_returns.loc[top_20_all_data.index][columns]
    min_variance_portfolio_data[year] = [top_20_monthly_returns, pf_weights]


for key in min_variance_portfolio_data.keys():
    pf_returns = min_variance_portfolio_data[key][0]
    pf_weights = min_variance_portfolio_data[key][1]

    pf_returns['weights'] = pf_weights
    pf_returns.loc['Portfolio return ln'] = np.dot(pf_weights, pf_returns)

    pf_returns.loc['Portfolio return'] = pf_returns.loc['Portfolio return ln'].apply(lambda x: np.exp(x))

    pf_returns.loc['Portfolio return ln']['weights'] = np.NaN
    pf_returns.loc['Portfolio return']['weights'] = np.NaN

    # pf_returns.to_excel(f'results/min_var_portfolio/min_var_{key}.xlsx')  # todo uncomment to save results to excel

# top_100_tot_returns.to_excel('interim_results/tot_returns.xlsx')  # todo uncomment to save results to excel
# df_percentile.to_excel('interim_results/percentiles_all.xlsx')  # todo uncomment to save results to excel
# df_group.to_excel('interim_results/groups_all.xlsx')  # todo uncomment to save results to excel
