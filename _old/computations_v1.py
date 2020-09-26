
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.formula.translate import Translator

wb = openpyxl.load_workbook('data/Switzerland_vfinal_modified.xlsx')

ws = wb['Sheet2']   # sheet name                Todo needs to be checked before running the program
block_height = 12   # number of rows per ticker Todo needs to be checked before running the program
max_col = 'EB'      # last column               Todo needs to be checked before running the program
row = 5             # start row (first ticker)  Todo needs to be checked before running the program
free_space = 1      # empty room between blocks Todo needs to be checked before running the program

company_tickers = []
raw_parameters = ['Total Return Index', 'Pension Deficit', 'Market Capitalization', 'Pension Liabilities',
                  'Service Cost', 'Interest Cost', 'Pension Paid', 'EBIT', 'Asset Allocation Debt',
                  'Asset Allocation Stocks']
parameters = ['R1', 'R2', 'R3', 'R4', 'R5']


for company_index in range(0, 100):
    print()
    print('Currently on company: {}'.format(company_index))

    # hide the two rows needed only for computation of Pension Deficit
    ws.row_dimensions.group(row+2, row+3, hidden=False)

    # Pension Deficit
    ws.insert_rows(row+2)
    ws.cell(row=row + 2, column=1).value = 'Pension Deficit'
    ws.cell(row=row + 2, column=2).value = 'CUSTOM_COMPUTATION'

    formula_str = "=C{}-C{}".format(row + 3, row + 4)
    first_col_location = 'C{}'.format(row + 2)
    ws[first_col_location] = formula_str

    for column in range(4, 100):
        ws.cell(row=row + 2, column=column).value = \
            Translator(formula_str,
                       origin=first_col_location).translate_formula('{}{}'.format(get_column_letter(column), row + 2))

    # move to next company
    row += block_height + free_space + 1  # + x depends on the number of rows inserted

wb.save('data/Switzerland_vfinal_added_computation.xlsx')
