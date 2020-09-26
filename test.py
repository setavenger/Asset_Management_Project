import openpyxl
import numpy as np
import pandas as pd
from util import try_float_iter, try_float, even_groups
import math
import datetime

wb = openpyxl.load_workbook('data/Switzerland_vfinal_modified.xlsx')

ws = wb['Sheet2']   # sheet name                Todo needs to be checked before running the program
block_height = 12   # number of rows per ticker Todo needs to be checked before running the program
row = 5             # start row (first ticker)  Todo needs to be checked before running the program
free_space = 1      # empty room between blocks Todo needs to be checked before running the program

df_total_returns = pd.DataFrame(dtype='float64', columns=[datetime.date(2020, 6, 30)] +
                                                         [ws.cell(row=row, column=i).value.date()
                                                          for i in range(4, 99)])

for company_index in range(0, 100):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = f'{ws.cell(row=row, column=1).value}'

    # 3DAABF; 75BF6E; E3E0D9;

    total_returns = np.array(list(try_float_iter([ws.cell(row=row + 1, column=i).value for i in range(3, 100)])))

    df_total_returns.loc[company_ticker] = [math.log(total_returns[i] / total_returns[i + 1])
                                            for i in range(len(total_returns) - 1)]

    # move to next company
    row += block_height + free_space + 0  # + x depends on the number of rows inserted
