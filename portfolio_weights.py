import openpyxl
import numpy as np
import pandas as pd
from util import try_float_iter, try_float, even_groups
import math

wb = openpyxl.load_workbook('data/Switzerland_vfinal_modified.xlsx')

ws = wb['Sheet2']   # sheet name                Todo needs to be checked before running the program
block_height = 12   # number of rows per ticker Todo needs to be checked before running the program
max_col = 'EB'      # last column               Todo needs to be checked before running the program
row = 5             # start row (first ticker)  Todo needs to be checked before running the program
free_space = 1      # empty room between blocks Todo needs to be checked before running the program

years_obv = ['2013', '2014', '2015', '2016', '2017', '2018', '2019', '2020']

r1 = pd.DataFrame(dtype='float64', columns=pd.MultiIndex.from_product([years_obv, ['r1', 'Total_Return']]))
r2 = pd.DataFrame(dtype='float64', columns=pd.MultiIndex.from_product([years_obv, ['r2', 'Total_Return']]))
r3 = pd.DataFrame(dtype='float64', columns=pd.MultiIndex.from_product([years_obv, ['r3', 'Total_Return']]))
r4 = pd.DataFrame(dtype='float64', columns=pd.MultiIndex.from_product([years_obv, ['r4', 'Total_Return']]))


total_return_2019 = pd.DataFrame(dtype='float64')

for company_index in range(0, 100):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = '{}'.format(ws.cell(row=row, column=1).value, company_index + 1)

    years_index_counter = 0
    month_shift = 0
    for year in reversed(years_obv):
        # !Hint! : currently only computes the first year further code needed to apply for whole time range
        # 3DAABF; 75BF6E; E3E0D9;

        projected_benefit_obligation = try_float(ws.cell(row=row + 2, column=month_shift + 21).value)
        fair_value_plan_assets = try_float(ws.cell(row=row + 3, column=month_shift + 21).value)
        market_cap = try_float(ws.cell(row=row + 4, column=month_shift + 21).value)
        service_cost = try_float(ws.cell(row=row + 6, column=month_shift + 21).value)
        interest_cost = try_float(ws.cell(row=row + 7, column=month_shift + 21).value)
        pension_paid = try_float(ws.cell(row=row + 8, column=month_shift + 21).value)
        ebit = try_float(ws.cell(row=row + 9, column=month_shift + 21).value)
        asset_allocation_debt = try_float(ws.cell(row=row + 10, column=month_shift + 21).value)
        asset_allocation_stocks = try_float(ws.cell(row=row + 11, column=month_shift + 21).value)

        total_return_raw = np.array(
            [ws.cell(row=row + 1, column=i).value for i in range(month_shift + 3, month_shift + 16)])

        total_return = np.array(list(try_float_iter(total_return_raw)))
        total_return_year = math.log(total_return[0] / total_return[-1])

        # todo eliminate nan and compute with remaining values
        pension_deficit = projected_benefit_obligation - fair_value_plan_assets
        r1.loc[company_ticker, year] = [pension_deficit / market_cap, total_return_year]
        r2.loc[company_ticker, year] = [projected_benefit_obligation / market_cap, total_return_year]
        r3.loc[company_ticker, year] = [(service_cost + interest_cost - pension_paid) / ebit, total_return_year]
        r4.loc[company_ticker, year] = [asset_allocation_debt - asset_allocation_stocks, total_return_year]

        month_shift += 12
        pass

    # move to next company
    row += block_height + free_space + 0  # + x depends on the number of rows inserted

# total_return_2019.to_excel('tests/ranked_test_v2.xlsx')

r1_2020 = r1.loc[:, '2020'].sort_values(ascending=True, by='r1').dropna()
r1_ranks_all = [(i / len(r1_2020['r1'])) for i in range(1, len(r1_2020['r1'])+1)]
r1_ranks = even_groups(len(r1_2020), 5, True)
r1_2020 = r1_2020.assign(r1_ranks=r1_ranks)
r1_2020 = r1_2020.assign(percentile=r1_ranks_all)
r1_2020_rank1 = r1_2020[r1_2020['r1_ranks'] == 1]
print('Average Annual Return r1-2020-group1:', r1_2020_rank1['Total_Return'].mean())


