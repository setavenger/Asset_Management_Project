df_main_bal.loc[('KOMN SW Equity', 'R4')]
df_main_bal.loc[('KOMN SW Equity', 'R5'), :] = 50
for year in years_obv[:-1]:
    for index in df_r5_mod.index:
        df_main_bal.loc[index, year] = df_r5.loc[index][year]

for index in df_r5_mod.index:
    df_main_bal.loc[index, :] = df_r5.loc[index]

for index in df_r5_mod.index:
    df_main_bal.loc[index, :] = df_r5_mod.loc[index]

df_r5_mod = pd.concat([df_r5], keys=['R5']).swaplevel()
for year in years_obv[:-1]:
    for index in df_r5_mod.index:
        df_main_bal.loc[index, year] = df_r5_mod.loc[index][year]

df_main_bal = df_main_bal.sort_index()
df_main_bal = df_main_bal.drop('2019')
df_main_bal = df_main_bal.drop('2019', 1)
[7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6].reverse()
reversed([7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6])
list(reversed([7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5, 6]))


def momentum_dates(year: int, df: pd.DataFrame) -> Tuple[list, list, list]:
    momentum_1_approved = []
    momentum_6_approved = []
    momentum_12_approved = []

    months = [6, 5, 4, 3, 2, 1, 12, 11, 10, 9, 8, 7]
    for i in range(len(months)):
        if i < 1:
            momentum_1_approved.append((year, months[i]))
        if i < 6:
            momentum_6_approved.append((year, months[i]))
            momentum_12_approved.append((year, months[i]))
        else:
            momentum_12_approved.append((year - 1, months[i]))

    columns_momentum_1 = []
    columns_momentum_6 = []
    columns_momentum_12 = []

    for i in df.columns:
        if (i.year, i.month) in momentum_1_approved:
            columns_momentum_1.append(i)
        if (i.year, i.month) in momentum_6_approved:
            columns_momentum_6.append(i)
        if (i.year, i.month) in momentum_12_approved:
            columns_momentum_12.append(i)

    return columns_momentum_1, columns_momentum_6, columns_momentum_12


def momentum_dates(year: int, df: pd.DataFrame):
    momentum_1_approved = []
    momentum_6_approved = []
    momentum_12_approved = []

    months = [6, 5, 4, 3, 2, 1, 12, 11, 10, 9, 8, 7]
    for i in range(len(months)):
        if i < 1:
            momentum_1_approved.append((year, months[i]))
        if i < 6:
            momentum_6_approved.append((year, months[i]))
            momentum_12_approved.append((year, months[i]))
        else:
            momentum_12_approved.append((year - 1, months[i]))

    columns_momentum_1 = []
    columns_momentum_6 = []
    columns_momentum_12 = []

    for i in df.columns:
        if (i.year, i.month) in momentum_1_approved:
            columns_momentum_1.append(i)
        if (i.year, i.month) in momentum_6_approved:
            columns_momentum_6.append(i)
        if (i.year, i.month) in momentum_12_approved:
            columns_momentum_12.append(i)

    return columns_momentum_1, columns_momentum_6, columns_momentum_12


momentum_dates(2018, df_total_returns)
columns_momentum_12


def momentum_dates(year: int, df: pd.DataFrame) -> Tuple[list, list, list]:
    momentum_1_approved = []
    momentum_6_approved = []
    momentum_12_approved = []

    months = [6, 5, 4, 3, 2, 1, 12, 11, 10, 9, 8, 7]
    for i in range(len(months)):
        if i < 1:
            momentum_1_approved.append((year + 1, months[i]))
        if i < 6:
            momentum_6_approved.append((year + 1, months[i]))
            momentum_12_approved.append((year + 1, months[i]))
        else:
            momentum_12_approved.append((year, months[i]))

    columns_momentum_1 = []
    columns_momentum_6 = []
    columns_momentum_12 = []

    for i in df.columns:
        if (i.year, i.month) in momentum_1_approved:
            columns_momentum_1.append(i)
        if (i.year, i.month) in momentum_6_approved:
            columns_momentum_6.append(i)
        if (i.year, i.month) in momentum_12_approved:
            columns_momentum_12.append(i)

    return columns_momentum_1, columns_momentum_6, columns_momentum_12


def momentum_dates(year: int, df: pd.DataFrame):
    momentum_1_approved = []
    momentum_6_approved = []
    momentum_12_approved = []

    months = [6, 5, 4, 3, 2, 1, 12, 11, 10, 9, 8, 7]
    for i in range(len(months)):
        if i < 1:
            momentum_1_approved.append((year + 1, months[i]))
        if i < 6:
            momentum_6_approved.append((year + 1, months[i]))
            momentum_12_approved.append((year + 1, months[i]))
        else:
            momentum_12_approved.append((year, months[i]))

    columns_momentum_1 = []
    columns_momentum_6 = []
    columns_momentum_12 = []

    for i in df.columns:
        if (i.year, i.month) in momentum_1_approved:
            columns_momentum_1.append(i)
        if (i.year, i.month) in momentum_6_approved:
            columns_momentum_6.append(i)
        if (i.year, i.month) in momentum_12_approved:
            columns_momentum_12.append(i)

    return columns_momentum_1, columns_momentum_6, columns_momentum_12


columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(2018, df_total_returns)
columns_momentum_1
columns_momentum_6
df_total_returns[columns_momentum_1] += 1
df_total_returns[columns_momentum_1].apply(lambda x: x - 2)
df_total_returns[columns_momentum_1]
columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(int(year), df_total_returns)
df_total_returns[columns_momentum_1].apply(+1)
df_total_returns[columns_momentum_1].apply(lambda x: x + 1)
mom_1_2018 = df_total_returns[columns_momentum_1].apply(lambda x: x + 1)
mom_1_2018.product(axis=1)
mom_1_2018['Momentum 1'] = mom_1_2018.product(axis=1)
mom_6_2018['Momentum 6'] = mom_6_2018.product(axis=1)
mom_6_2018['Momentum 6'] = mom_6_2018['Momentum 6'] - 1
mom_6_2018 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
mom_6_2018['Momentum 6'] = mom_6_2018.product(axis=1) - 1
df_regression_data = df_main_bal.copy()

# !Hint! : add R5 to coefficient table
for year in years_obv[:-1]:
    columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(int(year), df_total_returns)

    momentum_1 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_1['Momentum 1'] = momentum_1.product(axis=1) - 1

    momentum_6 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_6['Momentum 6'] = momentum_6.product(axis=1) - 1

    momentum_12 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_12['Momentum 12'] = momentum_12.product(axis=1) - 1

    for index in df_total_returns.index:
        df_regression_data.loc[(index, 'Momentum 1'), year] = momentum_1['Momentum 1'].loc[index]
        df_regression_data.loc[(index, 'Momentum 6'), year] = momentum_1['Momentum 6'].loc[index]
        df_regression_data.loc[(index, 'Momentum 12'), year] = momentum_1['Momentum 12'].loc[index]

df_regression_data = df_main_bal.copy()

# !Hint! : add R5 to coefficient table
for year in years_obv[:-1]:
    columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(int(year), df_total_returns)

    momentum_1 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_1['Momentum 1'] = momentum_1.product(axis=1) - 1

    momentum_6 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_6['Momentum 6'] = momentum_6.product(axis=1) - 1

    momentum_12 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_12['Momentum 12'] = momentum_12.product(axis=1) - 1

    for index in df_total_returns.index:
        df_regression_data.loc[(index, 'Momentum 1'), year] = momentum_1['Momentum 1'].loc[index]
        df_regression_data.loc[(index, 'Momentum 6'), year] = momentum_6['Momentum 6'].loc[index]
        df_regression_data.loc[(index, 'Momentum 12'), year] = momentum_12['Momentum 12'].loc[index]

df_regression_data.sort_index()
df_regression_data = df_main_bal.copy()

# !Hint! : add R5 to coefficient table
for year in years_obv[:-1]:
    columns_momentum_1, columns_momentum_6, columns_momentum_12 = momentum_dates(int(year), df_total_returns)
    columns_return_year = build_date_list(int(year), df_total_returns)

    momentum_1 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_1['Momentum 1'] = momentum_1.product(axis=1) - 1

    momentum_6 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_6['Momentum 6'] = momentum_6.product(axis=1) - 1

    momentum_12 = df_total_returns[columns_momentum_6].apply(lambda x: x + 1)
    momentum_12['Momentum 12'] = momentum_12.product(axis=1) - 1

    return_year = df_total_returns[columns_return_year].apply(lambda x: x + 1)
    return_year['Return Year'] = return_year.product(axis=1) - 1

    for index in df_total_returns.index:
        df_regression_data.loc[(index, 'Momentum 1'), year] = momentum_1['Momentum 1'].loc[index]
        df_regression_data.loc[(index, 'Momentum 6'), year] = momentum_6['Momentum 6'].loc[index]
        df_regression_data.loc[(index, 'Momentum 12'), year] = momentum_12['Momentum 12'].loc[index]
        df_regression_data.loc[(index, 'Return Year'), year] = return_year['Return Year'].loc[index]

df_regression_data = df_regression_data.sort_index()
import math

math.exp(0.0640307539352384)
lambda x: math.exp(x), [0, 1, 2]
list(lambda x: math.exp(x), [0, 1, 2])
list(map(lambda x: math.exp(x), [0, 1, 2]))
df_regression_data.to_excel('/regression_data.xlsx')
df_regression_data.to_excel('results/regression_data.xlsx')
df_regression_data['2018'].transpose()
df_regression_data['2018']
df_regression_data.transpose()
test = df_regression_data['2018']
for row in test[:20]:
    print(row)

keys
for key in keys:
    df_regression_data.loc[key]

for key in keys:
    print(df_regression_data.loc[key])

for key in keys:
    print(df_regression_data.loc[key]['2018'])

df_regression_data_raw.iloc[0]
df_regression_data_raw.iloc[1]
df_regression_data_raw.iloc['ABBN SW Equity']
df_regression_data_raw.loc['ABBN SW Equity']
df_regression_data_raw.loc['ABBN SW Equity', '2018']
df_regression_data_raw.loc['ABBN SW Equity', '2018'].transpose()
type(df_regression_data_raw.loc['ABBN SW Equity', '2018'].transpose())
df_regression_data_raw.loc['ABBN SW Equity', '2018'].T
df_regression_data_raw.loc['ABBN SW Equity', '2018'].values
vals = df_regression_data_raw.loc['ABBN SW Equity', '2018'].values
vals.transpose()
a = np.array([1, 2, 3, 4])
a.transpose()
a = np.array([[1, 2, 3, 4]])
a.T
df_regression_data_raw.loc['ABBN SW Equity', '2018'].to_frame().transpose()
pd.concat([df_regression_data_raw.loc['ABBN SW Equity', '2018'].to_frame().transpose()], keys=['ABBN SW Equity'])
pd.concat([df_regression_data_raw.loc['ABBN SW Equity', '2018'].to_frame().transpose()],
          keys=['ABBN SW Equity']).columns
for year in years_obv[:-1]:
    for company in keys:
        print(f'Transposing {year} {company}')
        row = pd.concat([df_regression_data_raw.loc[company, year].to_frame().transpose()],
                        keys=[company])
        df_regression_data = df_regression_data.append(row)

df_regression_data.to_excel('results/regression_data.xlsx', merge_cells=False)
df_regression_data[:'Return Year']
df_regression_data.loc[:, :'Return Year']
df_regression_data.loc[:, :'R5']
import statsmodels.api as sm

X = df_regression_data.loc[:, :'R5']
y = df_regression_data['Return Year']

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
model.summary()
import statsmodels.api as sm

X = df_regression_data.loc[:, :'R5'].dropna()
y = df_regression_data['Return Year'].dropna()

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
model.summary()
import statsmodels.api as sm

df_regression_data = df_regression_data.dropna()
X = df_regression_data.loc[:, :'R5']
y = df_regression_data['Return Year']

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
model.summary()
import statsmodels.api as sm

df_regression_data = df_regression_data.dropna()
X = df_regression_data.loc[:, :'R5']
y = df_regression_data['Return Year']
X = sm.add_constant(X)

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
print(model.summary())
from sklearn import linear_model

lm = linear_model.LinearRegression()
model = lm.fit(X, Y)
lm = linear_model.LinearRegression()
model = lm.fit(X, y)
lm.score(X, y)
import statsmodels.api as sm

df_regression_data = df_regression_data.dropna()
X = df_regression_data.loc['R5']
y = df_regression_data['Return Year']
X = sm.add_constant(X)

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
print(model.summary())
import statsmodels.api as sm

df_regression_data = df_regression_data.dropna()
X = df_regression_data['R5']
y = df_regression_data['Return Year']
X = sm.add_constant(X)

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
print(model.summary())
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures

model.score()
model.score(x_poly, y)
polynomial_features = PolynomialFeatures(degree=2)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
polynomial_features = PolynomialFeatures(degree=4)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=3)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=10)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=7)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=6)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=1)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=5)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
y_poly_pred = model.predict(x_poly)

rmse = np.sqrt(mean_squared_error(y, y_poly_pred))
r2 = r2_score(y, y_poly_pred)
r2
polynomial_features = PolynomialFeatures(degree=3)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
y_poly_pred = model.predict(x_poly)

rmse = np.sqrt(mean_squared_error(y, y_poly_pred))
r2 = r2_score(y, y_poly_pred)
rmse
polynomial_features = PolynomialFeatures(degree=3)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = sm.OLS(y, x_poly).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
model.summary()
polynomial_features = PolynomialFeatures(degree=3)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = sm.OLS(y, x_poly).fit()
predictions = model.predict(x_poly)  # make the predictions by the model

# Print out the statistics
model.summary()
model.summary()
polynomial_features = PolynomialFeatures(degree=4)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=5.5)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
polynomial_features = PolynomialFeatures(degree=5)
x_poly = polynomial_features.fit_transform(df_regression_data.loc[:, :'R5'])

model = LinearRegression()
model.fit(x_poly, y)
print(model.score(x_poly, y))
import statsmodels.api as sm

df_regression_data_clean = df_regression_data.dropna()
X = df_regression_data_clean.loc[:, :'R5']
y = df_regression_data_clean['Return Year']

# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
print(model.summary())
import statsmodels.api as sm

df_regression_data_clean = df_regression_data.dropna()
X = df_regression_data_clean.loc[:, :'R5']
y = df_regression_data_clean['Return Year']
X = sm.add_constant(X)
# Note the difference in argument order
model = sm.OLS(y, X).fit()
predictions = model.predict(X)  # make the predictions by the model

# Print out the statistics
print(model.summary())
model.summary2()
predictions
predictions.index
pd.MultiIndex.from_tuples(predictions.index, names=('company', 'year'))
predictions.index = pd.MultiIndex.from_tuples(predictions.index, names=('company', 'year'))
predictions.xs('2018', level=1).sort_values()
predictions.xs('2018', level=1).sort_values()[:20]
predictions.xs('2018', level=1).sort_values()[:-20]
predictions.xs('2018', level=1).sort_values('descending')[:20]
predictions.xs('2018', level=1).sort_values(ascending=False)[:20]
return_year['Std. Dev.'] = return_year.std(axis=1)
min_variance_data = pd.DataFrame(columns=pd.MultiIndex.from_product([years_obv[:-1], ['Return Year', 'Std. Dev.']]))
min_variance_data[('2018', 'Return Year')] = return_year['Return Year']
for year in years_obv[:-1]:
    top_20 = predictions.xs(year, level=1).sort_values(ascending=False)[:20]

top_20
min_variance_data.loc[top_20.index][('2018', 'Std. Dev.')]
min_variance_data.loc['IMPN SW Equity'][('2018', 'Std. Dev.')]
min_variance_data.loc['TXGN SW Equity'][('2018', 'Std. Dev.')]
top_20_std = min_variance_data.loc[top_20.index][('2018', 'Std. Dev.')]
top_20.loc[:, idx[:, 'Std. Dev.']]
top_20.loc[:, pd.idx[:, 'Std. Dev.']]
top_20.loc[:, pd.IndexSlice[:, 'Std. Dev.']]
top_20 = top_20.to_frame('Expected Return')
top_20.join(min_variance_data[('2018', 'Std. Dev.')])
top_20.join(min_variance_data[('2018', 'Std. Dev.')]).values
weights = np.array([[0.5], [0.5]])
l = np.ones((1, 2))
l
l = np.ones((2, 1))
weigths, l
weights, l
np.dot(weights.T, l)


def objective_function(weights, l):
    return np.dot(weights, l)


bounds = [(0, None) for _ in len(weigths)]
bounds = [(0, None) for _ in len(weights)]
bounds = [(0, None) for i in len(weights)]
weights.T
df_total_returns.corr()
df_total_returns.T.corr()
df_total_returns.T.cov()
df_total_returns.loc[top_20.index].T.cov()
np.empty((1, 20)).fill(0.05)
w = np.empty((1, 20)).fill(0.05)
w = np.empty((1, 20))
w.fill(0.05)
w.T
np.dot(w.T, df_total_returns.loc[top_20.index].T.cov().values, w)
np.dot(w.T, df_total_returns.loc[top_20.index].T.cov().values)
np.dot(w, df_total_returns.loc[top_20.index].T.cov().values, w.T)
np.dot(w, df_total_returns.loc[top_20.index].T.cov().values)
np.dot(np.dot(w, df_total_returns.loc[top_20.index].T.cov().values), w.T)
std_dev = [0.2, 0.15]
np.dot(np.dot(weights, std_dev), weights)
std_dev = [[0.2, 0.15]]
np.dot(np.dot(weights.T, std_dev), weights)
np.dot(np.dot(weights, std_dev), weights.T)
np.dot(weights, std_dev)
one = np.dot(weights, std_dev)
np.dot(one, weights.T)
np.dot(weights.T, one)
np.zeros((2, 2))
np.dot(np.dot(weights, np.zeros((2, 2))), weights.T)
np.dot(np.dot(weights.T, np.zeros((2, 2))), weights)
std_dev = np.array(std_dev)
np.dot(std_dev, std_dev.T)
np.dot(std_dev.T, std_dev)
omega = np.dot(std_dev.T, std_dev)
np.dot(np.dot(weights.T, omega, weights))
np.dot(np.dot(weights.T, omega), weights)
(weights.T, omega, weights)
np.dot(weights.T, omega, weights)
np.dot(weights.T, omega)
one = np.dot(weights.T, omega)
np.dot(one, weights)
weights
np.sqrt(np.dot(one, weights))
/ Users / setor / Downloads / Mean
Variance
Diagramm.xlsx
import openpyxl

wb = openpyxl.load_workbook('/Users/setor/Downloads/Mean Variance Diagramm.xlsx')
ws = wb['Sheet 1']
w = [ws.cell(row=9, column=i).value.date() for i in range(1, 6)]
w = [ws.cell(row=9, column=i).value for i in range(1, 6)]
w
cov = []
for row in range(1, 6):
    cov.append([ws.cell(row=row, column=i).value for i in range(1, 6)])

import numpy as np

w = np.array(w)
cov = np.array(cov)
np.sqrt(np.dot(np.dot(w, co), w))
bounds = [(0, None) for i in weights]
bounds = [(0, None) for i in w]


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov)))


constraint = LinearConstraint(np.ones(len(w)), lb=1, ub=1)
res = minimize(
    objective_function,
    x0=10 * np.random.random(n_buyers),
    args=(prices,),
    constraints=constraint,
    bounds=bounds, )

res = minimize(
    objective_function,
    x0=10 * np.random.random(10),
    args=(cov,),
    constraints=constraint,
    bounds=bounds, )

res = minimize(
    objective_function,
    x0=5 * np.random.random(10),
    args=(cov,),
    constraints=constraint,
    bounds=bounds, )

sum(res.x)
np.sqrt(np.dot(np.dot(w.T, cov), w))
np.sqrt(np.dot(np.dot(w, cov), w))
np.sqrt(np.dot(np.dot(w, cov), w.T))
np.sqrt(np.dot(np.dot(res.x, cov), res.x))
wb = openpyxl.load_workbook('/Users/setor/Downloads/Mean Variance Diagramm.xlsx')
ws = wb['Sheet1']
exp_ret = [ws.cell(row=11, column=i).value for i in range(1, 6)]
exp_ret = np.array(exp_ret)
exp_ret = np.array([exp_ret])
w_1 = np.array([res.x])
np.dot(w_1, exp_ret)
np.dot(w_1.T, exp_ret)
sum(np.dot(w_1.T, exp_ret))
np.dot(w_1, exp_ret.T)
cons = [constraint,
        {'type': 'eq', 'fun': con_real}]


def objective_function(weights, cove):
    return np.sqrt(np.dot(np.dot(weights, cove), weights.T))


res = minimize(
    objective_function,
    x0=5 * np.random.random(5),
    args=(cov,),
    constraints=constraint,
    bounds=bounds, )


def objective_function(cov, weights):
    return np.sqrt(np.dot(np.dot(weights, cove), weights.T))


def objective_function(cov, weights):
    return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


objective_function(cov, w)


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


res = minimize(
    objective_function,
    x0=[0, 0, 0, 0.9, 0],
    args=(cov,),
    constraints=constraint,
    bounds=bounds, )

res = minimize(
    objective_function,
    x0=[0, 0, 0, 0.9, 0],
    args=(cov, exp_ret),
    constraints=constraint,
    bounds=bounds, )

res = minimize(
    objective_function,
    x0=[0, 0, 0, 0.9, 0],
    args=(cov,),
    constraints=constraint,
    bounds=bounds,
)


def target_return(w):
    return np.dot(w, exp_ret.T)


def target_return(w):
    return np.dot(w, exp_ret.T) - 0.12


bounds
bounds = [(0, 1) for i in w]
res = minimize(
    objective_function,
    x0=[0.2, 0.2, 0.2, 0.2, 0.2],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

res.x


def target_return(w):
    return np.dot(w, exp_ret.T) - 0.15


bounds = [(0.005, 1) for i in w]
res = minimize(
    objective_function,
    x0=[0.2, 0.2, 0.2, 0.2, 0.2],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
bounds = [(0.05, 1) for i in w]
bounds = [(0.1, 1) for i in w]
res = minimize(
    objective_function,
    x0=[1, 1, 1, 1, 1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
res = minimize(
    objective_function,
    x0=[10, 1, 1, 1, 1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
res = minimize(
    objective_function,
    x0=[0.9, 0.9, 0.9, 1, 1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
res = minimize(
    objective_function,
    x0=[0.1, 0.9, 0.9, 1, 1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
bounds = [(0.1, 0.3) for i in w]
res = minimize(
    objective_function,
    x0=[0.1, 0.4, 0.2, 0.2, 0.1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)
res
res = minimize(
    objective_function,
    x0=[0.01, 0.4, 0.2, 0.2, 0.1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res.x)


def target_return(w):
    return np.dot(w, top_20_all_data['Expected Return'].T) - 0.08


def target_return(w):
    return np.dot(w, exp_ret.T) - 0.08


res = minimize(
    objective_function,
    x0=[0.01, 0.4, 0.2, 0.2, 0.1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
print(res)
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]
res = minimize(
    objective_function,
    x0=[0.01, 0.4, 0.2, 0.2, 0.1],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

print(res)
res.success
res = minimize(
    objective_function,
    x0=np.array([0.005 for _ in range(len(w))]),
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
res = minimize(
    objective_function,
    x0=np.array([[0.005 for _ in range(len(w))]]),
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

top_20_all_data['Expected Return'].T.shape()
top_20_all_data['Expected Return'].T
top_20_all_data['Expected Return'].shape()
top_20_all_data['Expected Return'].shape
top_20_all_data['Expected Return'].T.shape
top_20_all_data['Expected Return'].T.values.shape
[top_20_all_data['Expected Return'].T.values].shape
filtered['2020-06-30']
filtered.iloc[:, 1]
filtered.iloc[:, 1].to_excel('ticker_list')
filtered.iloc[:, 1].to_excel('ticker_list.xlsx')
w_init = [0.05] * 20
bounds = [(0.005, 0.2) for i in w_init]  # min and max weight in min var pf => no overweight and no omitted shares


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.02


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov)))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init,
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)
from scipy.optimize import minimize, LinearConstraint

w_init = [0.05] * 20
bounds = [(0.005, 0.2) for i in w_init]  # min and max weight in min var pf => no overweight and no omitted shares


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.02


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov)))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init,
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

w_init = np.full((20, 1), 0.05)
bounds = [(0.005, 0.2) for i in w_init]  # min and max weight in min var pf => no overweight and no omitted shares


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.02


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov)))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init,
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

bounds = np.array(
    [[(0.005, 0.2) for i in w_init]])  # min and max weight in min var pf => no overweight and no omitted shares


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.02


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init,
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

np.array([top_20_all_data['Expected Return'].T.values]).shape
w.shape


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.18


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init,
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

print(res)
len(w_init)
np.full((0, 20), 0.05)
np.full((2, 20), 0.05)
np.full((1, 20), 0.05)
np.full((1, 20), 0.05).shape


def target_return(w):
    return np.dot(w, np.array([top_20_all_data['Expected Return'].T.values])) - 0.18


def objective_function(weights, cov):
    return np.sqrt(np.dot(np.dot(weights, cov), weights.T))


constraint = LinearConstraint(np.ones(len(w_init)), lb=1, ub=1)  # sum of weight must 1
cons = [constraint,
        {'type': 'eq', 'fun': target_return}]

res = minimize(
    objective_function,
    x0=w_init[0],
    args=(cov,),
    constraints=cons,
    bounds=bounds,
)

print(res)
w_init[0]
np.array([top_20_all_data['Expected Return'].T.values])
np.dot(w_init, np.array([top_20_all_data['Expected Return'].values]))
np.dot(w_init, np.array([top_20_all_data['Expected Return'].T.values]))
np.dot(w_init, np.array([top_20_all_data['Expected Return'].values]).T)
np.sqrt(np.dot(np.dot(w_init, cova), w_init.T))
np.sqrt(np.dot(np.dot(w_init, cov), w_init.T))
np.sqrt(np.dot(np.dot(w_init.T, cov), w_init))
w_init
w_init.shape
np.dot(w_init.T, cov)
np.dot(w_init, cov)
cov
cov.values
np.sqrt(np.dot(np.dot(w_init, cov.values), w_init.T))
runfile('/Users/setor/PycharmProjects/Asset_Management_Project/computations.py',
        wdir='/Users/setor/PycharmProjects/Asset_Management_Project')