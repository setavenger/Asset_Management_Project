import openpyxl
import pandas as pd
import numpy as np
import datetime
from util import try_float_iter, even_groups, build_date_list
import math

wb = openpyxl.load_workbook('data/Switzerland_vfinal_clean.xlsx')

# Todo needs to be checked before running the program
ws_balance_sheet = wb['balance_sheet_data']  # sheet name
ws_total_return = wb['total_returns']  # sheet name
ws_exclusions = wb['exclusions']  # sheet name

free_space = 1  # empty room between blocks
# !Hint! : removed 2020 only moving until 2019 Balance Sheet Data
years_obv = ['2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019']
risk_factors_all = ['R1', 'R2', 'R3', 'R4', 'R5']
# days to keep 30.06.2011
first_day = datetime.datetime(2011, 6, 30)  # 30.06.2011

company_blacklist = []
row_ticker = 2
while True:
    if not ws_exclusions.cell(row=row_ticker, column=1).value:
        break
    company_blacklist.append(ws_exclusions.cell(row=row_ticker, column=1).value)
    row_ticker += 1

row_tot_return = 4

df_total_returns = pd.DataFrame(dtype='float64', columns=[datetime.date(2020, 6, 30)] +
                                                         [ws_total_return.cell(row=row_tot_return,
                                                                               column=i).value.date()
                                                          for i in range(4, 110)])

for company_index in range(0, 203):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = f'{ws_total_return.cell(row=row_tot_return, column=1).value}'

    total_returns = np.array(list(try_float_iter(
        [ws_total_return.cell(row=row_tot_return + 1, column=i).value for i in range(3, 111)])))

    df_total_returns.loc[company_ticker] = [math.log(total_returns[i] / total_returns[i + 1])
                                            for i in range(len(total_returns) - 1)]

    # move to next company
    row_tot_return += 2 + free_space  # + x depends on the number of rows inserted

filtered = df_total_returns[~df_total_returns.index.isin(company_blacklist)]
df_total_returns = filtered.iloc[:100]

# !Hint! : balance sheet data from here
#  plan is to create single dfs per company and concat them via loc
#  this will create a MultiIndex
#  the columns will be matched in case a company misses data for a specific year

df_balance_sheet_data = pd.DataFrame(columns=years_obv)

frames = []
keys = []
row_bal = 4

# !Hint! : NESN first year balance ebit missing

for company_index in range(0, 203):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_ticker = f'{ws_balance_sheet.cell(row=row_bal, column=1).value}'
    if company_ticker in company_blacklist:
        row_bal += 12 + free_space  # + x depends on the number of rows inserted
        continue
    keys.append(company_ticker)
    columns_df = ['2019']
    column = 4
    while True:
        if ws_balance_sheet.cell(row=row_bal, column=column).value is None:
            break
            pass
        columns_df.append(str(ws_balance_sheet.cell(row=row_bal, column=column).value.year))
        column += 1

    df_company = pd.DataFrame(columns=columns_df)

    # !Hint! : the company risk calculations and df
    projected_benefit_obligation = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 2, column=i).value
                                                                 for i in range(3, 12)])))
    fair_value_plan_assets = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 3, column=i).value
                                                           for i in range(3, 12)])))
    market_cap = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 4, column=i).value
                                               for i in range(3, 12)])))
    service_cost = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 6, column=i).value
                                                 for i in range(3, 12)])))
    interest_cost = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 7, column=i).value
                                                  for i in range(3, 12)])))
    pension_paid = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 8, column=i).value
                                                 for i in range(3, 12)])))
    ebit = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 9, column=i).value
                                         for i in range(3, 12)])))
    asset_allocation_debt = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 10, column=i).value
                                                          for i in range(3, 12)])))
    asset_allocation_stocks = np.array(list(try_float_iter([ws_balance_sheet.cell(row=row_bal + 11, column=i).value
                                                            for i in range(3, 12)])))

    pension_deficit = projected_benefit_obligation - fair_value_plan_assets

    r1 = pension_deficit / market_cap
    r2 = projected_benefit_obligation / market_cap
    r3 = (service_cost + interest_cost - pension_paid) / ebit
    r4 = asset_allocation_debt - asset_allocation_stocks

    df_company.loc['R1'] = r1[:len(columns_df)]
    df_company.loc['R2'] = r2[:len(columns_df)]
    df_company.loc['R3'] = r3[:len(columns_df)]
    df_company.loc['R4'] = r4[:len(columns_df)]

    frames.append(df_company)

    # move to next company
    row_bal += 12 + free_space  # + x depends on the number of rows inserted

df_main_bal = pd.concat(frames, keys=keys)

# !Hint! : only the top 100
df_main_bal = df_main_bal[:400]  # 100 companies à 4 risks
keys = keys[:100]
# !Hint! : from here on the quintiles will be created
df_group = pd.DataFrame(columns=years_obv[:-1], index=df_main_bal.index)
df_percentile = pd.DataFrame(columns=years_obv[:-1], index=df_main_bal.index)

for year in years_obv[:-1]:
    for risk_factor in ['R1', 'R2', 'R3', 'R4']:
        print(f'{year}-{risk_factor}')
        sub_df = df_main_bal.xs(risk_factor, level=1)[year].to_frame(f'{year}_{risk_factor}')
        sub_df['percentile'] = sub_df[f'{year}_{risk_factor}'].rank(pct=True)
        sub_df = sub_df.sort_values(ascending=True, by=f'{year}_{risk_factor}').dropna()
        group = even_groups(len(sub_df), 5, True)
        sub_df = sub_df.assign(group=group)
        # todo check whether all are lowest to highest risk increasing
        sub_df_group = pd.concat([sub_df], keys=[risk_factor]).swaplevel()['group']
        sub_df_percentile = pd.concat([sub_df], keys=[risk_factor]).swaplevel()['percentile']

        # df_group
        for index in sub_df_group.index:
            df_group.loc[index][year] = sub_df_group.loc[index]
        # df_percentile
        for index in sub_df_group.index:
            df_percentile.loc[index][year] = sub_df_percentile.loc[index]

# !Hint! : R5 calculations
df_r5 = pd.DataFrame(columns=years_obv[:-1])
for comp in keys:
    df_r5.loc[comp] = df_percentile.loc[comp].mean()

for year in df_r5.columns:
    df = df_r5[year].sort_values(ascending=True).dropna()
    group = even_groups(len(df), 5, True)
    df = df.to_frame(year).assign(group=group)
    df['percentile'] = df[f'{year}'].rank(pct=True)
    for comp in df.index:
        df_percentile.loc[(comp, 'R5'), year] = df['percentile'].loc[comp]
        df_percentile = df_percentile.sort_index()

for year in years_obv[:-1]:
    print(f'R5-{year}')
    df = df_r5[year]
    df = df.sort_values(ascending=True).dropna()
    group = even_groups(len(df), 5, True)
    df = df.to_frame('R5_Percentile').assign(group=group)
    df_group_r5 = pd.concat([df], keys=['R5']).swaplevel()['group']
    for index in df_group_r5.index:
        df_group.loc[index, year] = df_group_r5.loc[index]

# top_100_tot_returns.to_excel('interim_results/tot_returns.xlsx')
# df_percentile.to_excel('interim_results/percentiles_all.xlsx')
# df_group.to_excel('interim_results/groups_all.xlsx')

# !Hint! : create total returns for all quintile portfolios

total_return_dfs = {}
quartiles_only = {}
for year in years_obv[:-1]:
    columns_new = build_date_list(int(year), df_total_returns)
    risk_dfs = []
    df_quartiles_only = pd.DataFrame(index=pd.MultiIndex.from_product([risk_factors_all,
                                                                       [f'Q-{rank}' for rank in range(1, 6)]]),
                                     columns=columns_new)
    for risk in risk_factors_all:
        rank_dfs = []
        for rank in range(1, 6):
            print(f'Total Return {year}--{risk} Q-{rank}')
            indices = df_group.xs(risk, level=1)[year][(df_group.xs(risk, level=1)[year] == rank)].index
            df = df_total_returns.loc[indices][df_total_returns.columns.intersection(columns_new)]
            df_quartiles_only.loc[risk, f'Q-{rank}'] = df.mean()
            rank_dfs.append(df)

        rank_df = pd.concat(rank_dfs, keys=[f'Q-{i}' for i in range(1, 6)])
        risk_dfs.append(rank_df)

    df_for_year = pd.concat(risk_dfs, keys=risk_factors_all)

    total_return_dfs[year] = df_for_year
    df_quartiles_only['mean'] = df_quartiles_only.mean(axis=1)
    df_quartiles_only['std'] = df_quartiles_only.std(axis=1)
    quartiles_only[year] = df_quartiles_only
