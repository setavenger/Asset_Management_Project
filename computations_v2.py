
import openpyxl
import numpy as np
import pandas as pd
from util import try_float_iter, try_date

wb = openpyxl.load_workbook('data/Switzerland_vfinal_modified.xlsx')

ws = wb['Sheet2']   # sheet name                Todo needs to be checked before running the program
block_height = 12   # number of rows per ticker Todo needs to be checked before running the program
max_col = 'EB'      # last column               Todo needs to be checked before running the program
row = 5             # start row (first ticker)  Todo needs to be checked before running the program
free_space = 1      # empty room between blocks Todo needs to be checked before running the program


raw_parameters = ['Total Return Index', 'Pension Deficit', 'Market Capitalization', 'Pension Liabilities',
                  'Service Cost', 'Interest Cost', 'Pension Paid', 'EBIT', 'Asset Allocation Debt',
                  'Asset Allocation Stocks']

parameters = ['R1', 'R2', 'R3', 'R4', 'R5']
parameters_2 = ['R1', 'R2', 'R3', 'R4']
company_tickers = []
data = np.empty((0, 97), float)
dates = np.array(list(try_date([ws.cell(row=row, column=i).value for i in range(3, 100)])))

for company_index in range(0, 100):
    print()
    print('Currently on company: {}'.format(company_index))

    # Add Ticker to company list
    company_tickers.append('{}'.format(ws.cell(row=row, column=1).value, company_index+1))

    projected_benefit_obligation_raw = np.array([ws.cell(row=row+2, column=i).value for i in range(3, 100)])
    fair_value_plan_assets_raw = np.array([ws.cell(row=row+3, column=i).value for i in range(3, 100)])
    market_cap_raw = np.array([ws.cell(row=row + 4, column=i).value for i in range(3, 100)])
    service_cost_raw = np.array([ws.cell(row=row + 6, column=i).value for i in range(3, 100)])
    interest_cost_raw = np.array([ws.cell(row=row + 7, column=i).value for i in range(3, 100)])
    pension_paid_raw = np.array([ws.cell(row=row + 8, column=i).value for i in range(3, 100)])
    ebit_raw = np.array([ws.cell(row=row + 9, column=i).value for i in range(3, 100)])
    asset_allocation_debt_raw = np.array([ws.cell(row=row + 10, column=i).value for i in range(3, 100)])
    asset_allocation_stocks_raw = np.array([ws.cell(row=row + 11, column=i).value for i in range(3, 100)])

    # convert strings to floats for computation purposes
    projected_benefit_obligation = np.array(list(try_float_iter(projected_benefit_obligation_raw)))
    fair_value_plan_assets = np.array(list(try_float_iter(fair_value_plan_assets_raw)))
    market_cap = np.array(list(try_float_iter(market_cap_raw)))
    service_cost = np.array(list(try_float_iter(service_cost_raw)))
    interest_cost = np.array(list(try_float_iter(interest_cost_raw)))
    pension_paid = np.array(list(try_float_iter(pension_paid_raw)))
    ebit = np.array(list(try_float_iter(ebit_raw)))
    asset_allocation_debt = np.array(list(try_float_iter(asset_allocation_debt_raw)))
    asset_allocation_stocks = np.array(list(try_float_iter(asset_allocation_stocks_raw)))

    pension_deficit = np.subtract(projected_benefit_obligation, fair_value_plan_assets)
    r1 = pension_deficit / market_cap
    r2 = projected_benefit_obligation / market_cap
    r3 = (service_cost + interest_cost - pension_paid) / ebit
    r4 = asset_allocation_debt - asset_allocation_stocks

    data = np.append(data, np.array([np.array(r1)]), axis=0)
    data = np.append(data, np.array([np.array(r2)]), axis=0)
    data = np.append(data, np.array([np.array(r3)]), axis=0)
    data = np.append(data, np.array([np.array(r4)]), axis=0)

    # fill for r5
    # data = np.append(data, np.full((1, 97), np.NaN), axis=0) Todo add when r5 is included but not computed

    # move to next company
    row += block_height + free_space + 0  # + x depends on the number of rows inserted


multi_idx = pd.MultiIndex.from_product([company_tickers, parameters_2])
company_data = pd.DataFrame(data=data, index=multi_idx, columns=dates)

company_data.to_excel('data/output_v3.xlsx')
