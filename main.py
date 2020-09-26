import openpyxl
from datetime import datetime
from openpyxl.utils.cell import get_column_letter
from util import validate_day


wb = openpyxl.load_workbook('data/Switzerland100 v1.xlsx')

ws = wb['Sheet2']   # sheet name                Todo needs to be checked before running the program
block_height = 12   # number of rows per ticker Todo needs to be checked before running the program
max_col = 'EB'      # last column               Todo needs to be checked before running the program
row = 5             # start row (first ticker)  Todo needs to be checked before running the program
free_space = 1      # empty room between blocks Todo needs to be checked before running the program
# days to keep 30.06.2012
first_day = datetime(2012, 6, 30)  # 30.06.2012 Todo needs to be checked before running the program


for company_index in range(0, 100):
    print()
    print('Currently on company: {}'.format(company_index))
    circuitBreaker = True
    column = 4
    deleted_counter = 0
    # delete all weekends (double dates / non-trading days)
    while circuitBreaker:

        if ws.cell(row=row, column=column).value is None:
            print(row, column)
            print('Deleted {} dates on company {}'.format(deleted_counter, company_index))
            row += block_height + free_space
            break

        if not validate_day(ws.cell(row=row, column=column).value, first_day):  # Todo check: formula was changed untested
            deleted_counter += 1
            for i in range(block_height + 1):
                ws.cell(row=row + i, column=column).value = None
            # after deletion concatenate
            range_str = "{}{}:EB{}".format(get_column_letter(column + 1), row, row + block_height)
            ws.move_range(range_str, cols=-1)
        column += 1

wb.save('data/Switzerland_vfinal_modified.xlsx')
